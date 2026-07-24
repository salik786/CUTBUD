#!/usr/bin/env python3
"""
Sync the AI Hairstyle Library spreadsheet into the `style_catalog` table.

Metadata only — this script never touches imageUrl/leftImageUrl/rightImageUrl/
backImageUrl/displayAngle. Images are generated manually in Higgsfield (using
the front/left/right/back prompts this script syncs in) and uploaded through
the admin panel's existing image uploader, which shows those prompts inline.

Usage:
    python3 scripts/sync_hairstyles.py --file "/path/to/workbook.xlsx" --limit 3   # test on first 3 rows
    python3 scripts/sync_hairstyles.py --file "/path/to/workbook.xlsx"             # full run
    python3 scripts/sync_hairstyles.py --file "/path/to/workbook.xlsx" --deactivate-existing
        # one-time: set active=false on every CURRENT row before syncing, so
        # old hand-entered styles don't sit alongside the spreadsheet's ones
        # under a different id. Safe to run once; a no-op on later runs.

Requires DIRECT_URL (or DATABASE_URL as a fallback) in the environment —
loaded from .env in the project root if present. Never hardcode credentials.
"""

import argparse
import os
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

PROJECT_ROOT = Path(__file__).resolve().parent.parent


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


def get_db_url() -> str:
    load_dotenv(PROJECT_ROOT / ".env")
    url = os.environ.get("DIRECT_URL") or os.environ.get("DATABASE_URL")
    if not url:
        sys.exit("DIRECT_URL or DATABASE_URL must be set (checked environment and .env)")
    return url


def split_pipe(value) -> str:
    """Spreadsheet multi-value cells are pipe-separated; the DB uses commas."""
    if not value:
        return ""
    parts = [p.strip() for p in str(value).split("|") if p.strip()]
    return ",".join(parts)


def build_face_shape_map(wb) -> dict:
    ws = wb["Hairstyle_FaceShapes"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    by_id = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        if not r.get("hairstyle_id") or not r.get("face_shape"):
            continue
        shape = str(r["face_shape"]).strip()
        if shape not in by_id[r["hairstyle_id"]]:
            by_id[r["hairstyle_id"]].append(shape)
    return {hid: ",".join(shapes) for hid, shapes in by_id.items()}


def build_prompts_map(wb) -> dict:
    ws = wb["Prompts"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    by_id = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        hid = r.get("hairstyle_id")
        if not hid:
            continue
        by_id[hid] = {
            "frontPrompt": r.get("front_prompt") or None,
            "leftPrompt": r.get("left_prompt") or None,
            "rightPrompt": r.get("right_prompt") or None,
            "backPrompt": r.get("back_prompt") or None,
            "tryonPrompt": r.get("tryon_prompt") or None,
        }
    return by_id


def make_description(r: dict) -> str:
    parts = []
    if r.get("category"):
        parts.append(f"{r['category']} style")
    if r.get("hair_length"):
        parts.append(f"{r['hair_length'].lower()} length")
    if r.get("maintenance_level"):
        parts.append(f"{r['maintenance_level'].lower()} maintenance")
    base = ", ".join(parts) if parts else r.get("name", "Hairstyle")
    inspiration = f" {r['inspiration']}." if r.get("inspiration") else "."
    return f"{base}.{inspiration}".strip()


def make_base_prompt(r: dict) -> str:
    bits = [r.get("name") or "Hairstyle"]
    if r.get("hair_length"):
        bits.append(f"{r['hair_length'].lower()} length")
    if r.get("category"):
        bits.append(f"{r['category'].lower()} category")
    beard = r.get("beard_pairing")
    if beard and beard != "None":
        bits.append(f"beard pairing: {beard.lower()}")
    return ", ".join(bits)


def build_row(r: dict, face_shapes: dict, prompts: dict) -> dict:
    hid = r["hairstyle_id"]
    status = (r.get("status") or "").strip().lower()
    data = {
        "id": hid,
        "name": r.get("name") or hid,
        "description": make_description(r),
        "basePrompt": make_base_prompt(r),
        "faceShapeFit": face_shapes.get(hid) or "oval",  # required column; "oval" only if the sheet truly has no mapping
        "category": r.get("category"),
        "textureCompat": split_pipe(r.get("suitable_hair_texture")),
        "density": split_pipe(r.get("hair_density_fit")),
        "lengthCategory": r.get("hair_length"),
        "maintenance": r.get("maintenance_level"),
        "beardPairing": r.get("beard_pairing"),
        "occasion": split_pipe(r.get("occasion_tags")),
        "targetAudience": split_pipe(r.get("target_audience")),
        "inspiredBy": r.get("inspiration"),
        "modelId": r.get("model_id"),
        "trendScore": int(r["trend_score"]) if r.get("trend_score") not in (None, "") else None,
        "active": status == "active",
    }
    data.update(prompts.get(hid, {}))
    # Empty-string multi-value fields (nothing after the pipe-split) should be
    # NULL, not "", so they don't show up as an empty-but-"set" filter value.
    for key in ("textureCompat", "density", "occasion", "targetAudience"):
        if data[key] == "":
            data[key] = None
    return data


# Columns that are ALWAYS refreshed from the spreadsheet on re-run (it's the
# source of truth for these). description/basePrompt are deliberately
# excluded — they're synthesized once on first insert and then left alone so
# manual edits made later in the admin panel survive a re-sync. Image columns
# are never touched by this script at all.
UPDATE_COLUMNS = [
    "name",
    "faceShapeFit",
    "category",
    "textureCompat",
    "density",
    "lengthCategory",
    "maintenance",
    "beardPairing",
    "occasion",
    "targetAudience",
    "inspiredBy",
    "modelId",
    "trendScore",
    "active",
    "frontPrompt",
    "leftPrompt",
    "rightPrompt",
    "backPrompt",
    "tryonPrompt",
]
INSERT_COLUMNS = ["id", "description", "basePrompt"] + UPDATE_COLUMNS


def upsert_row(cur, data: dict) -> None:
    cols = INSERT_COLUMNS
    values = [data[c] for c in cols]
    quoted_cols = ", ".join(f'"{c}"' for c in cols)
    placeholders = ", ".join(["%s"] * len(cols))
    set_clause = ", ".join(f'"{c}" = EXCLUDED."{c}"' for c in UPDATE_COLUMNS)
    sql = f"""
        INSERT INTO style_catalog ({quoted_cols})
        VALUES ({placeholders})
        ON CONFLICT (id) DO UPDATE SET {set_clause}
    """
    cur.execute(sql, values)


def deactivate_existing(conn) -> int:
    with conn.cursor() as cur:
        cur.execute("UPDATE style_catalog SET active = false")
        count = cur.rowcount
    conn.commit()
    return count


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", required=True, help="Path to the .xlsx workbook")
    parser.add_argument("--limit", type=int, default=None, help="Only process the first N rows (for testing)")
    parser.add_argument(
        "--deactivate-existing",
        action="store_true",
        help="Set active=false on every current row before syncing (one-time cleanup)",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.file).expanduser()
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    db_url = get_db_url()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    face_shapes = build_face_shape_map(wb)
    prompts = build_prompts_map(wb)

    ws = wb["Hairstyles"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [
        dict(zip(headers, row))
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[0]
    ]
    if args.limit:
        rows = rows[: args.limit]

    print(f"Loaded {len(rows)} hairstyle row(s) from {xlsx_path.name}")

    conn = psycopg2.connect(db_url)
    conn.autocommit = False

    if args.deactivate_existing:
        n = deactivate_existing(conn)
        print(f"Deactivated {n} existing row(s) (active=false) before syncing.")

    succeeded, failed = [], []
    for r in rows:
        hid = r.get("hairstyle_id", "<unknown>")
        try:
            data = build_row(r, face_shapes, prompts)
            with conn.cursor() as cur:
                upsert_row(cur, data)
            conn.commit()
            succeeded.append(hid)
            print(f"  ok   {hid}  {data['name']}")
        except Exception as e:  # noqa: BLE001 — intentionally broad: keep the batch going
            conn.rollback()
            failed.append((hid, str(e)))
            print(f"  FAIL {hid}: {e}")

    conn.close()

    print(f"\nDone: {len(succeeded)} succeeded, {len(failed)} failed.")
    if failed:
        print("Failed rows:")
        for hid, err in failed:
            print(f"  {hid}: {err}")
        sys.exit(1)


if __name__ == "__main__":
    main()
